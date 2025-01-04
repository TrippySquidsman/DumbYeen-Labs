import requests
import openpyxl
import warnings
from urllib3.exceptions import InsecureRequestWarning

# Suppress the warning
warnings.simplefilter("ignore", InsecureRequestWarning)

# Clear the screen
print("\033[H\033[J")

# Proxmox API details
base_url = "https://10.0.0.1:8006/api2/json"
username = "root@pam"
password = "byHhHNIg072qh&hM@r4kT8O0"

# Filepath and Worksheet Details
excel_path = "https://d.docs.live.net/E88524C5BB44C92A/Documents/Homelab/Homelab%20Service%20Overview.xlsx"
hostname_column = "Hostname"
pve_id_column = "PVE ID"

# Create or load an Excel file
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# Identify the columns
for col_idx, col in enumerate(sheet[1], start=1):  # Assuming row 1 is the header row
    if col.value == hostname_column:
        hostname_col_idx = col_idx
        print(f"Hostname column found at index: {hostname_col_idx}")
    elif col.value == pve_id_column:
        pve_id_col_idx = col_idx
        print(f"PVE ID column found at index: {pve_id_col_idx}")

# Load the hostnames from the Excel file into a list
hostnames = [row[hostname_col_idx - 1].value for row in sheet.iter_rows(min_row=2)]  # Skip header row

# Print the hostnames into a table and mention the row number
for idx, hostname in enumerate(hostnames, start=2):
    print(f"Row {idx}: {hostname}, current PVE ID: {sheet.cell(row=idx, column=pve_id_col_idx).value}")

# Get a ticket (login)
response = requests.post(f"{base_url}/access/ticket", data={
    "username": username,
    "password": password
}, verify=False)  # Disable SSL verification for simplicity

response_data = response.json()
ticket = response_data["data"]["ticket"]
csrf_token = response_data["data"]["CSRFPreventionToken"]

# Fetch container details
headers = {"CSRFPreventionToken": csrf_token}
cookies = {"PVEAuthCookie": ticket}

# Fetch container details for host01 to host05
hosts = ["host01", "host02", "host03", "host04", "host05"]
container_data = {}
for host in hosts:
    containers = requests.get(f"{base_url}/nodes/{host}/lxc", headers=headers, cookies=cookies, verify=False).json()
    for container in containers["data"]:
        container_data[container["name"]] = container["vmid"]

# Write container names and IDs to Excel
for container_name, container_id in container_data.items():
    for row in sheet.iter_rows(min_row=2):  # Skip header row
        hostname = row[hostname_col_idx - 1].value
        if hostname in container_name:
            row[pve_id_col_idx - 1].value = container_data[hostname]
            print(f"{container_name} found at row {row[0].row}")
            print(f" Updated ID to {container_id}")

# Save the updated Excel file
workbook.save(excel_path)